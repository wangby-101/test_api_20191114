# -*- coding: utf-8 -*-
# @Time    : 2019/11/15 14:39
# @Author  : Wang Bingyin
# @Email   : xxxxxxx@163.com
# @File    : do_excel.py
# @Software: PyCharm
import openpyxl
from Common.base_path import data_file
class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expect = None
        self.actual = None
        self.result = None

class Do_Excel:

    def __init__(self, fielname):
        try:
            self.filename = fielname
            self.workbook = openpyxl.load_workbook(filename=fielname)
        except FileNotFoundError as e:
            print("{} file not found!".format(fielname))
            raise e

    def get_case(self, sheetname):
        sheet = self.workbook[sheetname]
        max_row = sheet.max_row
        cases = []
        for row in range(2, max_row+1):
           case = Case()
           case.case_id = sheet.cell(row, column=1).value
           case.title = sheet.cell(row, column=2).value
           case.url = sheet.cell(row, column=3).value
           case.method = sheet.cell(row, column=5).value
           case.data = sheet.cell(row, column=4).value
           case.expected = sheet.cell(row, column=6).value
           cases.append(case)
        return cases

    def write_result(self, sheetname, case_id, actual, result):
        sheet = self.workbook[sheetname]
        max_row = sheet.max_row
        for row in range(2, max_row+1):
            if sheet.cell(row, 1).value == case_id:
                sheet.cell(row, column=7).value = actual
                sheet.cell(row, column=8).value = result
                self.workbook.save(self.filename)
                break

if __name__ == '__main__':
    de = Do_Excel(data_file)
    cases = de.get_case("login")
    for case in cases:
        de.write_result(sheetname='login', case_id=case.case_id, actual=case.url, result=False)